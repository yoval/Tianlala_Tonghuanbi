#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
import numpy as np


# In[2]:


import time
now = time.strftime('%Y-%m-%d')


# In[3]:


fliePath = r'C:\Users\admin\OneDrive\甜啦啦\周月报\2023年05月第1&2周\汇总表2023.5.1~5.14.xlsx'


# 以下通过总表制作同环比分表

# In[4]:


df_zongbiao = pd.read_excel(fliePath,sheet_name="总表",header = 3)
#df_zongbiao[["大区经理", "省经理","区域经理"]] = df_zongbiao[["大区经理", "省经理","区域经理"]].replace(np.nan, "无")
df_zongbiao[["大区经理", "省经理","区域经理"]] = df_zongbiao[["大区经理", "省经理","区域经理"]].replace(np.nan, "错误，请提醒我重做")


# In[5]:


def Tonghuanbi(a,b):
    c = (a-b)/b
    return c


# 全量店铺-同比 ↓

# In[6]:


#大区经理-全量店铺-同比
values = ['本期是否营业','同比期是否营业','本期实收金额','同比期实收金额','本期堂食实收','同比期堂食实收',
                                                '本期外卖实收','同比期外卖实收','本期自提实收','同比期自提实收']
df_daqu_quanliang = df_zongbiao.pivot_table(index=['大区经理'],values = values,aggfunc=np.sum)


# In[7]:


df_daqu_quanliang['店铺增减'] = df_daqu_quanliang['本期是否营业']- df_daqu_quanliang['同比期是否营业']
#实收
df_daqu_quanliang['实收同比'] = Tonghuanbi(df_daqu_quanliang['本期实收金额'] ,df_daqu_quanliang['同比期实收金额'])
df_daqu_quanliang.sort_values("实收同比",inplace=True,ascending=False) #排序
#堂食实收
df_daqu_quanliang['堂食实收同比'] = Tonghuanbi(df_daqu_quanliang['本期堂食实收'] ,df_daqu_quanliang['同比期堂食实收'])
#外卖实收
df_daqu_quanliang['外卖实收同比'] = Tonghuanbi(df_daqu_quanliang['本期外卖实收'] ,df_daqu_quanliang['同比期外卖实收'])
#外卖实收
df_daqu_quanliang['自提实收同比'] = Tonghuanbi(df_daqu_quanliang['本期自提实收'] ,df_daqu_quanliang['同比期自提实收'])


# In[8]:


values = ['本期是否营业','同比期是否营业','店铺增减','本期实收金额','同比期实收金额','实收同比','本期堂食实收','同比期堂食实收',
          '堂食实收同比','本期外卖实收','同比期外卖实收','外卖实收同比','本期自提实收','同比期自提实收','自提实收同比']
df_daqu_quanliang = df_daqu_quanliang.reindex(columns=values)
df_daqu_quanliang.rename(columns={'本期是否营业': '本期店铺数','同比期是否营业': '同比期店铺数'}, inplace=True)


# In[9]:


total = df_daqu_quanliang[['本期店铺数', '同比期店铺数','店铺增减','本期实收金额','同比期实收金额',
            '本期堂食实收','同比期堂食实收','本期外卖实收','同比期外卖实收','本期自提实收','同比期自提实收']].sum()
total['实收同比'] = Tonghuanbi(total['本期实收金额'] , total['同比期实收金额'])

total['堂食实收同比'] = Tonghuanbi(total['本期堂食实收'] , total['同比期堂食实收'])
total['外卖实收同比'] = Tonghuanbi(total['本期外卖实收'] , total['同比期外卖实收'])
total['自提实收同比'] = Tonghuanbi(total['本期自提实收'] , total['同比期自提实收'])
total_df = pd.DataFrame(total).T.rename(index={0: '合计'})
df_daqu_quanliang = pd.concat([df_daqu_quanliang, total_df])


# In[10]:


df_daqu_quanliang.to_excel('同环比分表_%s.xlsx'%now, sheet_name='大区经理全量店铺同比')


# In[11]:


#省经理-全量店铺-同比
values = ['本期是否营业','同比期是否营业','本期实收金额','同比期实收金额','本期堂食实收','同比期堂食实收',
                                                '本期外卖实收','同比期外卖实收','本期自提实收','同比期自提实收']
df_daqu_quanliang = df_zongbiao.pivot_table(index=['省经理'],values = values,aggfunc=np.sum)


# In[12]:


df_daqu_quanliang['店铺增减'] = df_daqu_quanliang['本期是否营业']-df_daqu_quanliang['同比期是否营业']
#实收
df_daqu_quanliang['实收同比'] = (df_daqu_quanliang['本期实收金额'] - df_daqu_quanliang['同比期实收金额'])/df_daqu_quanliang['同比期实收金额']
df_daqu_quanliang.sort_values("实收同比",inplace=True,ascending=False) #排序
#堂食实收
df_daqu_quanliang['堂食实收同比'] = (df_daqu_quanliang['本期堂食实收'] - df_daqu_quanliang['同比期堂食实收'])/df_daqu_quanliang['同比期堂食实收']
#外卖实收
df_daqu_quanliang['外卖实收同比'] = (df_daqu_quanliang['本期外卖实收'] - df_daqu_quanliang['同比期外卖实收'])/df_daqu_quanliang['同比期外卖实收']
#外卖实收
df_daqu_quanliang['自提实收同比'] = (df_daqu_quanliang['本期自提实收'] - df_daqu_quanliang['同比期自提实收'])/df_daqu_quanliang['同比期自提实收']


# In[13]:


values = ['本期是否营业','同比期是否营业','店铺增减','本期实收金额','同比期实收金额','实收同比','本期堂食实收','同比期堂食实收',
          '堂食实收同比','本期外卖实收','同比期外卖实收','外卖实收同比','本期自提实收','同比期自提实收','自提实收同比']
df_daqu_quanliang = df_daqu_quanliang.reindex(columns=values)
df_daqu_quanliang.rename(columns={'本期是否营业': '本期店铺数','同比期是否营业': '同比期店铺数'}, inplace=True)


# In[14]:


total = df_daqu_quanliang[['本期店铺数', '同比期店铺数','店铺增减','本期实收金额','同比期实收金额',
            '本期堂食实收','同比期堂食实收','本期外卖实收','同比期外卖实收','本期自提实收','同比期自提实收']].sum()
total['实收同比'] = Tonghuanbi(total['本期实收金额'] , total['同比期实收金额'])

total['堂食实收同比'] = Tonghuanbi(total['本期堂食实收'] , total['同比期堂食实收'])
total['外卖实收同比'] = Tonghuanbi(total['本期外卖实收'] , total['同比期外卖实收'])
total['自提实收同比'] = Tonghuanbi(total['本期自提实收'] , total['同比期自提实收'])
total_df = pd.DataFrame(total).T.rename(index={0: '合计'})
df_daqu_quanliang = pd.concat([df_daqu_quanliang, total_df])


# In[15]:


with pd.ExcelWriter('同环比分表_%s.xlsx'%now, mode='a',engine="openpyxl") as writer:
    df_daqu_quanliang.to_excel(writer, sheet_name='省经理全量店铺同比')


# In[16]:


#区域经理-全量店铺-同比
values = ['本期是否营业','同比期是否营业','本期实收金额','同比期实收金额','本期堂食实收','同比期堂食实收',
                                                '本期外卖实收','同比期外卖实收','本期自提实收','同比期自提实收']
df_daqu_quanliang = df_zongbiao.pivot_table(index=['区域经理'],values = values,aggfunc=np.sum)


# In[17]:


df_daqu_quanliang['店铺增减'] = df_daqu_quanliang['本期是否营业']-df_daqu_quanliang['同比期是否营业']
#实收
df_daqu_quanliang['实收同比'] = (df_daqu_quanliang['本期实收金额'] - df_daqu_quanliang['同比期实收金额'])/df_daqu_quanliang['同比期实收金额']
df_daqu_quanliang.sort_values("实收同比",inplace=True,ascending=False) #排序
#堂食实收
df_daqu_quanliang['堂食实收同比'] = (df_daqu_quanliang['本期堂食实收'] - df_daqu_quanliang['同比期堂食实收'])/df_daqu_quanliang['同比期堂食实收']
#外卖实收
df_daqu_quanliang['外卖实收同比'] = (df_daqu_quanliang['本期外卖实收'] - df_daqu_quanliang['同比期外卖实收'])/df_daqu_quanliang['同比期外卖实收']
#外卖实收
df_daqu_quanliang['自提实收同比'] = (df_daqu_quanliang['本期自提实收'] - df_daqu_quanliang['同比期自提实收'])/df_daqu_quanliang['同比期自提实收']


# In[18]:


values = ['本期是否营业','同比期是否营业','店铺增减','本期实收金额','同比期实收金额','实收同比','本期堂食实收','同比期堂食实收',
          '堂食实收同比','本期外卖实收','同比期外卖实收','外卖实收同比','本期自提实收','同比期自提实收','自提实收同比']
df_daqu_quanliang = df_daqu_quanliang.reindex(columns=values)
df_daqu_quanliang.rename(columns={'本期是否营业': '本期店铺数','同比期是否营业': '同比期店铺数'}, inplace=True)


# In[19]:


total = df_daqu_quanliang[['本期店铺数', '同比期店铺数','店铺增减','本期实收金额','同比期实收金额',
            '本期堂食实收','同比期堂食实收','本期外卖实收','同比期外卖实收','本期自提实收','同比期自提实收']].sum()
total['实收同比'] = Tonghuanbi(total['本期实收金额'] , total['同比期实收金额'])

total['堂食实收同比'] = Tonghuanbi(total['本期堂食实收'] , total['同比期堂食实收'])
total['外卖实收同比'] = Tonghuanbi(total['本期外卖实收'] , total['同比期外卖实收'])
total['自提实收同比'] = Tonghuanbi(total['本期自提实收'] , total['同比期自提实收'])
total_df = pd.DataFrame(total).T.rename(index={0: '合计'})
df_daqu_quanliang = pd.concat([df_daqu_quanliang, total_df])


# In[20]:


with pd.ExcelWriter('同环比分表_%s.xlsx'%now, mode='a',engine="openpyxl") as writer:
    df_daqu_quanliang.to_excel(writer, sheet_name='区域经理全量店铺同比')


# 全量店铺-环比 ↓

# In[21]:


#大区经理-全量店铺-环比
values = ['本期是否营业','环比期是否营业','本期实收金额','环比期实收金额','本期堂食实收','环比期堂食实收',
                                                '本期外卖实收','环比期外卖实收','本期自提实收','环比期自提实收']
          
df_huanbi = df_zongbiao.pivot_table(index=['大区经理'],values = values,aggfunc=np.sum)


# In[22]:


df_huanbi['店铺增减'] = df_huanbi['本期是否营业']- df_huanbi['环比期是否营业']
#实收
df_huanbi['实收环比'] = (df_huanbi['本期实收金额'] - df_huanbi['环比期实收金额'])/df_huanbi['环比期实收金额']
df_huanbi.sort_values("实收环比",inplace=True,ascending=False) #排序
#堂食实收
df_huanbi['堂食实收环比'] = (df_huanbi['本期堂食实收'] - df_huanbi['环比期堂食实收'])/df_huanbi['环比期堂食实收']
#外卖实收
df_huanbi['外卖实收环比'] = (df_huanbi['本期外卖实收'] - df_huanbi['环比期外卖实收'])/df_huanbi['环比期外卖实收']
#外卖实收
df_huanbi['自提实收环比'] = (df_huanbi['本期自提实收'] - df_huanbi['环比期自提实收'])/df_huanbi['环比期自提实收']


# In[23]:


values = ['本期是否营业','环比期是否营业','店铺增减','本期实收金额','环比期实收金额','实收环比','本期堂食实收','环比期堂食实收',
          '堂食实收环比','本期外卖实收','环比期外卖实收','外卖实收环比','本期自提实收','环比期自提实收','自提实收环比']
df_huanbi = df_huanbi.reindex(columns=values)
df_huanbi.rename(columns={'本期是否营业': '本期店铺数','环比期是否营业': '环比期店铺数'}, inplace=True)


# In[24]:


total = df_huanbi[['本期店铺数', '环比期店铺数','店铺增减','本期实收金额','环比期实收金额',
            '本期堂食实收','环比期堂食实收','本期外卖实收','环比期外卖实收','本期自提实收','环比期自提实收']].sum()
total['实收环比'] = Tonghuanbi(total['本期实收金额'] , total['环比期实收金额'])

total['堂食实收环比'] = Tonghuanbi(total['本期堂食实收'] , total['环比期堂食实收'])
total['外卖实收环比'] = Tonghuanbi(total['本期外卖实收'] , total['环比期外卖实收'])
total['自提实收环比'] = Tonghuanbi(total['本期自提实收'] , total['环比期自提实收'])
total_df = pd.DataFrame(total).T.rename(index={0: '合计'})
df_huanbi = pd.concat([df_huanbi, total_df])


# In[25]:


with pd.ExcelWriter('同环比分表_%s.xlsx'%now, mode='a',engine="openpyxl") as writer:
    df_huanbi.to_excel(writer, sheet_name='大区经理全量店铺环比')


# In[26]:


#省经理-全量店铺-环比
values = ['本期是否营业','环比期是否营业','本期实收金额','环比期实收金额','本期堂食实收','环比期堂食实收',
                                                '本期外卖实收','环比期外卖实收','本期自提实收','环比期自提实收']
          
df_huanbi = df_zongbiao.pivot_table(index=['省经理'],values = values,aggfunc=np.sum)


# In[27]:


df_huanbi['店铺增减'] = df_huanbi['本期是否营业']- df_huanbi['环比期是否营业']
#实收
df_huanbi['实收环比'] = (df_huanbi['本期实收金额'] - df_huanbi['环比期实收金额'])/df_huanbi['环比期实收金额']
df_huanbi.sort_values("实收环比",inplace=True,ascending=False) #排序
#堂食实收
df_huanbi['堂食实收环比'] = (df_huanbi['本期堂食实收'] - df_huanbi['环比期堂食实收'])/df_huanbi['环比期堂食实收']
#外卖实收
df_huanbi['外卖实收环比'] = (df_huanbi['本期外卖实收'] - df_huanbi['环比期外卖实收'])/df_huanbi['环比期外卖实收']
#外卖实收
df_huanbi['自提实收环比'] = (df_huanbi['本期自提实收'] - df_huanbi['环比期自提实收'])/df_huanbi['环比期自提实收']


# In[28]:


values = ['本期是否营业','环比期是否营业','店铺增减','本期实收金额','环比期实收金额','实收环比','本期堂食实收','环比期堂食实收',
          '堂食实收环比','本期外卖实收','环比期外卖实收','外卖实收环比','本期自提实收','环比期自提实收','自提实收环比']
df_huanbi = df_huanbi.reindex(columns=values)
df_huanbi.rename(columns={'本期是否营业': '本期店铺数','环比期是否营业': '环比期店铺数'}, inplace=True)


# In[29]:


total = df_huanbi[['本期店铺数', '环比期店铺数','店铺增减','本期实收金额','环比期实收金额',
            '本期堂食实收','环比期堂食实收','本期外卖实收','环比期外卖实收','本期自提实收','环比期自提实收']].sum()
total['实收环比'] = Tonghuanbi(total['本期实收金额'] , total['环比期实收金额'])

total['堂食实收环比'] = Tonghuanbi(total['本期堂食实收'] , total['环比期堂食实收'])
total['外卖实收环比'] = Tonghuanbi(total['本期外卖实收'] , total['环比期外卖实收'])
total['自提实收环比'] = Tonghuanbi(total['本期自提实收'] , total['环比期自提实收'])
total_df = pd.DataFrame(total).T.rename(index={0: '合计'})
df_huanbi = pd.concat([df_huanbi, total_df])


# In[30]:


with pd.ExcelWriter('同环比分表_%s.xlsx'%now, mode='a',engine="openpyxl") as writer:
    df_huanbi.to_excel(writer, sheet_name='省经理全量店铺环比')


# In[31]:


#区域经理-全量店铺-环比
values = ['本期是否营业','环比期是否营业','本期实收金额','环比期实收金额','本期堂食实收','环比期堂食实收',
                                                '本期外卖实收','环比期外卖实收','本期自提实收','环比期自提实收']
          
df_huanbi = df_zongbiao.pivot_table(index=['区域经理'],values = values,aggfunc=np.sum)


# In[32]:


df_huanbi['店铺增减'] = df_huanbi['本期是否营业']- df_huanbi['环比期是否营业']
#实收
df_huanbi['实收环比'] = (df_huanbi['本期实收金额'] - df_huanbi['环比期实收金额'])/df_huanbi['环比期实收金额']
df_huanbi.sort_values("实收环比",inplace=True,ascending=False) #排序
#堂食实收
df_huanbi['堂食实收环比'] = (df_huanbi['本期堂食实收'] - df_huanbi['环比期堂食实收'])/df_huanbi['环比期堂食实收']
#外卖实收
df_huanbi['外卖实收环比'] = (df_huanbi['本期外卖实收'] - df_huanbi['环比期外卖实收'])/df_huanbi['环比期外卖实收']
#外卖实收
df_huanbi['自提实收环比'] = (df_huanbi['本期自提实收'] - df_huanbi['环比期自提实收'])/df_huanbi['环比期自提实收']


# In[33]:


values = ['本期是否营业','环比期是否营业','店铺增减','本期实收金额','环比期实收金额','实收环比','本期堂食实收','环比期堂食实收',
          '堂食实收环比','本期外卖实收','环比期外卖实收','外卖实收环比','本期自提实收','环比期自提实收','自提实收环比']
df_huanbi = df_huanbi.reindex(columns=values)
df_huanbi.rename(columns={'本期是否营业': '本期店铺数','环比期是否营业': '环比期店铺数'}, inplace=True)


# In[34]:


total = df_huanbi[['本期店铺数', '环比期店铺数','店铺增减','本期实收金额','环比期实收金额',
            '本期堂食实收','环比期堂食实收','本期外卖实收','环比期外卖实收','本期自提实收','环比期自提实收']].sum()
total['实收环比'] = Tonghuanbi(total['本期实收金额'] , total['环比期实收金额'])

total['堂食实收环比'] = Tonghuanbi(total['本期堂食实收'] , total['环比期堂食实收'])
total['外卖实收环比'] = Tonghuanbi(total['本期外卖实收'] , total['环比期外卖实收'])
total['自提实收环比'] = Tonghuanbi(total['本期自提实收'] , total['环比期自提实收'])
total_df = pd.DataFrame(total).T.rename(index={0: '合计'})
df_huanbi = pd.concat([df_huanbi, total_df])


# In[35]:


with pd.ExcelWriter('同环比分表_%s.xlsx'%now, mode='a',engine="openpyxl") as writer:
    df_huanbi.to_excel(writer, sheet_name='区域经理全量店铺环比')


# In[36]:


df_zongbiao.head(1)


# 存量店铺-同比

# In[37]:


#同比期存量店铺
df_cunliang = df_zongbiao[(df_zongbiao['本期是否营业']==1)]
pd.set_option('display.max_columns', None)


# In[38]:


df_cunliang.shape


# In[39]:


df_cunliang.head(1)


# In[40]:


df_cunliang = df_cunliang[(df_cunliang['同比期是否营业']==1)]


# In[41]:


df_cunliang.shape


# In[42]:


#大区经理-存量店铺-同比
values = ['本期是否营业','同比期是否营业','本期实收金额','同比期实收金额','本期堂食实收','同比期堂食实收',
                                                '本期外卖实收','同比期外卖实收','本期自提实收','同比期自提实收']
          
df_cunliang_pivot = df_cunliang.pivot_table(index=['大区经理'],values = values,aggfunc=np.sum)


# In[43]:


pd.set_option('display.max_columns', None)
df_cunliang_pivot.head()


# In[44]:


df_cunliang_pivot['店铺增减'] = df_cunliang_pivot['本期是否营业']-df_cunliang_pivot['同比期是否营业']
#实收
df_cunliang_pivot['实收同比'] = (df_cunliang_pivot['本期实收金额'] - df_cunliang_pivot['同比期实收金额'])/df_cunliang_pivot['同比期实收金额']
df_cunliang_pivot.sort_values("实收同比",inplace=True,ascending=False) #排序
#堂食实收
df_cunliang_pivot['堂食实收同比'] = (df_cunliang_pivot['本期堂食实收'] - df_cunliang_pivot['同比期堂食实收'])/df_cunliang_pivot['同比期堂食实收']
#外卖实收
df_cunliang_pivot['外卖实收同比'] = (df_cunliang_pivot['本期外卖实收'] - df_cunliang_pivot['同比期外卖实收'])/df_cunliang_pivot['同比期外卖实收']
#外卖实收
df_cunliang_pivot['自提实收同比'] = (df_cunliang_pivot['本期自提实收'] - df_cunliang_pivot['同比期自提实收'])/df_cunliang_pivot['同比期自提实收']


# In[45]:


values = ['本期是否营业','同比期是否营业','店铺增减','本期实收金额','同比期实收金额','实收同比','本期堂食实收','同比期堂食实收',
          '堂食实收同比','本期外卖实收','同比期外卖实收','外卖实收同比','本期自提实收','同比期自提实收','自提实收同比']
df_cunliang_pivot = df_cunliang_pivot.reindex(columns=values)
df_cunliang_pivot.rename(columns={'本期是否营业': '本期店铺数','同比期是否营业': '同比期店铺数'}, inplace=True)


# In[46]:


total = df_cunliang_pivot[['本期店铺数', '同比期店铺数','店铺增减','本期实收金额','同比期实收金额',
            '本期堂食实收','同比期堂食实收','本期外卖实收','同比期外卖实收','本期自提实收','同比期自提实收']].sum()
total['实收同比'] = Tonghuanbi(total['本期实收金额'] , total['同比期实收金额'])

total['堂食实收同比'] = Tonghuanbi(total['本期堂食实收'] , total['同比期堂食实收'])
total['外卖实收同比'] = Tonghuanbi(total['本期外卖实收'] , total['同比期外卖实收'])
total['自提实收同比'] = Tonghuanbi(total['本期自提实收'] , total['同比期自提实收'])
total_df = pd.DataFrame(total).T.rename(index={0: '合计'})
df_cunliang_pivot = pd.concat([df_cunliang_pivot, total_df])


# In[47]:


with pd.ExcelWriter('同环比分表_%s.xlsx'%now, mode='a',engine="openpyxl") as writer:
    df_cunliang_pivot.to_excel(writer, sheet_name='大区经理存量店铺同比')


# In[48]:


#省经理-存量店铺-同比
values = ['本期是否营业','同比期是否营业','本期实收金额','同比期实收金额','本期堂食实收','同比期堂食实收',
                                                '本期外卖实收','同比期外卖实收','本期自提实收','同比期自提实收']
          
df_cunliang_pivot = df_cunliang.pivot_table(index=['省经理'],values = values,aggfunc=np.sum)


# In[49]:


df_cunliang_pivot['店铺增减'] = df_cunliang_pivot['本期是否营业']-df_cunliang_pivot['同比期是否营业']
#实收
df_cunliang_pivot['实收同比'] = (df_cunliang_pivot['本期实收金额'] - df_cunliang_pivot['同比期实收金额'])/df_cunliang_pivot['同比期实收金额']
df_cunliang_pivot.sort_values("实收同比",inplace=True,ascending=False) #排序
#堂食实收
df_cunliang_pivot['堂食实收同比'] = (df_cunliang_pivot['本期堂食实收'] - df_cunliang_pivot['同比期堂食实收'])/df_cunliang_pivot['同比期堂食实收']
#外卖实收
df_cunliang_pivot['外卖实收同比'] = (df_cunliang_pivot['本期外卖实收'] - df_cunliang_pivot['同比期外卖实收'])/df_cunliang_pivot['同比期外卖实收']
#外卖实收
df_cunliang_pivot['自提实收同比'] = (df_cunliang_pivot['本期自提实收'] - df_cunliang_pivot['同比期自提实收'])/df_cunliang_pivot['同比期自提实收']


# In[50]:


values = ['本期是否营业','同比期是否营业','店铺增减','本期实收金额','同比期实收金额','实收同比','本期堂食实收','同比期堂食实收',
          '堂食实收同比','本期外卖实收','同比期外卖实收','外卖实收同比','本期自提实收','同比期自提实收','自提实收同比']
df_cunliang_pivot = df_cunliang_pivot.reindex(columns=values)
df_cunliang_pivot.rename(columns={'本期是否营业': '本期店铺数','同比期是否营业': '同比期店铺数'}, inplace=True)


# In[51]:


total = df_cunliang_pivot[['本期店铺数', '同比期店铺数','店铺增减','本期实收金额','同比期实收金额',
            '本期堂食实收','同比期堂食实收','本期外卖实收','同比期外卖实收','本期自提实收','同比期自提实收']].sum()
total['实收同比'] = Tonghuanbi(total['本期实收金额'] , total['同比期实收金额'])

total['堂食实收同比'] = Tonghuanbi(total['本期堂食实收'] , total['同比期堂食实收'])
total['外卖实收同比'] = Tonghuanbi(total['本期外卖实收'] , total['同比期外卖实收'])
total['自提实收同比'] = Tonghuanbi(total['本期自提实收'] , total['同比期自提实收'])
total_df = pd.DataFrame(total).T.rename(index={0: '合计'})
df_cunliang_pivot = pd.concat([df_cunliang_pivot, total_df])


# In[52]:


with pd.ExcelWriter('同环比分表_%s.xlsx'%now, mode='a',engine="openpyxl") as writer:
    df_cunliang_pivot.to_excel(writer, sheet_name='省经理存量店铺同比')


# In[53]:


#省经理-存量店铺-同比
values = ['本期是否营业','同比期是否营业','本期实收金额','同比期实收金额','本期堂食实收','同比期堂食实收',
                                                '本期外卖实收','同比期外卖实收','本期自提实收','同比期自提实收']
          
df_cunliang_pivot = df_cunliang.pivot_table(index=['区域经理'],values = values,aggfunc=np.sum)


# In[54]:


df_cunliang_pivot['店铺增减'] = df_cunliang_pivot['本期是否营业']-df_cunliang_pivot['同比期是否营业']
#实收
df_cunliang_pivot['实收同比'] = (df_cunliang_pivot['本期实收金额'] - df_cunliang_pivot['同比期实收金额'])/df_cunliang_pivot['同比期实收金额']
df_cunliang_pivot.sort_values("实收同比",inplace=True,ascending=False) #排序
#堂食实收
df_cunliang_pivot['堂食实收同比'] = (df_cunliang_pivot['本期堂食实收'] - df_cunliang_pivot['同比期堂食实收'])/df_cunliang_pivot['同比期堂食实收']
#外卖实收
df_cunliang_pivot['外卖实收同比'] = (df_cunliang_pivot['本期外卖实收'] - df_cunliang_pivot['同比期外卖实收'])/df_cunliang_pivot['同比期外卖实收']
#外卖实收
df_cunliang_pivot['自提实收同比'] = (df_cunliang_pivot['本期自提实收'] - df_cunliang_pivot['同比期自提实收'])/df_cunliang_pivot['同比期自提实收']


# In[55]:


values = ['本期是否营业','同比期是否营业','店铺增减','本期实收金额','同比期实收金额','实收同比','本期堂食实收','同比期堂食实收',
          '堂食实收同比','本期外卖实收','同比期外卖实收','外卖实收同比','本期自提实收','同比期自提实收','自提实收同比']
df_cunliang_pivot = df_cunliang_pivot.reindex(columns=values)
df_cunliang_pivot.rename(columns={'本期是否营业': '本期店铺数','同比期是否营业': '同比期店铺数'}, inplace=True)


# In[56]:


total = df_cunliang_pivot[['本期店铺数', '同比期店铺数','店铺增减','本期实收金额','同比期实收金额',
            '本期堂食实收','同比期堂食实收','本期外卖实收','同比期外卖实收','本期自提实收','同比期自提实收']].sum()
total['实收同比'] = Tonghuanbi(total['本期实收金额'] , total['同比期实收金额'])

total['堂食实收同比'] = Tonghuanbi(total['本期堂食实收'] , total['同比期堂食实收'])
total['外卖实收同比'] = Tonghuanbi(total['本期外卖实收'] , total['同比期外卖实收'])
total['自提实收同比'] = Tonghuanbi(total['本期自提实收'] , total['同比期自提实收'])
total_df = pd.DataFrame(total).T.rename(index={0: '合计'})
df_cunliang_pivot = pd.concat([df_cunliang_pivot, total_df])


# In[57]:


with pd.ExcelWriter('同环比分表_%s.xlsx'%now, mode='a',engine="openpyxl") as writer:
    df_cunliang_pivot.to_excel(writer, sheet_name='区域经理存量店铺同比')


# 存量店铺-环比

# In[58]:


#环比期存量店铺
df_cunliang = df_zongbiao[(df_zongbiao['本期是否营业']==1)]
df_cunliang = df_cunliang[(df_cunliang['环比期是否营业']==1)]


# In[59]:


#大区经理-存量店铺-环比
values = ['本期是否营业','环比期是否营业','本期实收金额','环比期实收金额','本期堂食实收','环比期堂食实收',
                                                '本期外卖实收','环比期外卖实收','本期自提实收','环比期自提实收']
          
df_cunliang_pivot = df_cunliang.pivot_table(index=['大区经理'],values = values,aggfunc=np.sum)


# In[60]:


df_cunliang_pivot['店铺增减'] = df_cunliang_pivot['本期是否营业']-df_cunliang_pivot['环比期是否营业']
#实收
df_cunliang_pivot['实收环比'] = (df_cunliang_pivot['本期实收金额'] - df_cunliang_pivot['环比期实收金额'])/df_cunliang_pivot['环比期实收金额']
df_cunliang_pivot.sort_values("实收环比",inplace=True,ascending=False) #排序
#堂食实收
df_cunliang_pivot['堂食实收环比'] = (df_cunliang_pivot['本期堂食实收'] - df_cunliang_pivot['环比期堂食实收'])/df_cunliang_pivot['环比期堂食实收']
#外卖实收
df_cunliang_pivot['外卖实收环比'] = (df_cunliang_pivot['本期外卖实收'] - df_cunliang_pivot['环比期外卖实收'])/df_cunliang_pivot['环比期外卖实收']
#外卖实收
df_cunliang_pivot['自提实收环比'] = (df_cunliang_pivot['本期自提实收'] - df_cunliang_pivot['环比期自提实收'])/df_cunliang_pivot['环比期自提实收']


# In[61]:


values = ['本期是否营业','环比期是否营业','店铺增减','本期实收金额','环比期实收金额','实收环比','本期堂食实收','环比期堂食实收',
          '堂食实收环比','本期外卖实收','环比期外卖实收','外卖实收环比','本期自提实收','环比期自提实收','自提实收环比']
df_cunliang_pivot = df_cunliang_pivot.reindex(columns=values)
df_cunliang_pivot.rename(columns={'本期是否营业': '本期店铺数','环比期是否营业': '环比期店铺数'}, inplace=True)


# In[62]:


total = df_cunliang_pivot[['本期店铺数', '环比期店铺数','店铺增减','本期实收金额','环比期实收金额',
            '本期堂食实收','环比期堂食实收','本期外卖实收','环比期外卖实收','本期自提实收','环比期自提实收']].sum()
total['实收环比'] = Tonghuanbi(total['本期实收金额'] , total['环比期实收金额'])

total['堂食实收环比'] = Tonghuanbi(total['本期堂食实收'] , total['环比期堂食实收'])
total['外卖实收环比'] = Tonghuanbi(total['本期外卖实收'] , total['环比期外卖实收'])
total['自提实收环比'] = Tonghuanbi(total['本期自提实收'] , total['环比期自提实收'])
total_df = pd.DataFrame(total).T.rename(index={0: '合计'})
df_cunliang_pivot = pd.concat([df_cunliang_pivot, total_df])


# In[63]:


with pd.ExcelWriter('同环比分表_%s.xlsx'%now, mode='a',engine="openpyxl") as writer:
    df_cunliang_pivot.to_excel(writer, sheet_name='大区经理存量店铺环比')


# In[64]:


#省经理-存量店铺-环比
values = ['本期是否营业','环比期是否营业','本期实收金额','环比期实收金额','本期堂食实收','环比期堂食实收',
                                                '本期外卖实收','环比期外卖实收','本期自提实收','环比期自提实收']
df_cunliang_pivot = df_cunliang.pivot_table(index=['省经理'],values = values,aggfunc=np.sum)


# In[65]:


df_cunliang_pivot['店铺增减'] = df_cunliang_pivot['本期是否营业']-df_cunliang_pivot['环比期是否营业']
#实收
df_cunliang_pivot['实收环比'] = (df_cunliang_pivot['本期实收金额'] - df_cunliang_pivot['环比期实收金额'])/df_cunliang_pivot['环比期实收金额']
df_cunliang_pivot.sort_values("实收环比",inplace=True,ascending=False) #排序
#堂食实收
df_cunliang_pivot['堂食实收环比'] = (df_cunliang_pivot['本期堂食实收'] - df_cunliang_pivot['环比期堂食实收'])/df_cunliang_pivot['环比期堂食实收']
#外卖实收
df_cunliang_pivot['外卖实收环比'] = (df_cunliang_pivot['本期外卖实收'] - df_cunliang_pivot['环比期外卖实收'])/df_cunliang_pivot['环比期外卖实收']
#外卖实收
df_cunliang_pivot['自提实收环比'] = (df_cunliang_pivot['本期自提实收'] - df_cunliang_pivot['环比期自提实收'])/df_cunliang_pivot['环比期自提实收']


# In[66]:


values = ['本期是否营业','环比期是否营业','店铺增减','本期实收金额','环比期实收金额','实收环比','本期堂食实收','环比期堂食实收',
          '堂食实收环比','本期外卖实收','环比期外卖实收','外卖实收环比','本期自提实收','环比期自提实收','自提实收环比']
df_cunliang_pivot = df_cunliang_pivot.reindex(columns=values)
df_cunliang_pivot.rename(columns={'本期是否营业': '本期店铺数','环比期是否营业': '环比期店铺数'}, inplace=True)


# In[67]:


total = df_cunliang_pivot[['本期店铺数', '环比期店铺数','店铺增减','本期实收金额','环比期实收金额',
            '本期堂食实收','环比期堂食实收','本期外卖实收','环比期外卖实收','本期自提实收','环比期自提实收']].sum()
total['实收环比'] = Tonghuanbi(total['本期实收金额'] , total['环比期实收金额'])

total['堂食实收环比'] = Tonghuanbi(total['本期堂食实收'] , total['环比期堂食实收'])
total['外卖实收环比'] = Tonghuanbi(total['本期外卖实收'] , total['环比期外卖实收'])
total['自提实收环比'] = Tonghuanbi(total['本期自提实收'] , total['环比期自提实收'])
total_df = pd.DataFrame(total).T.rename(index={0: '合计'})
df_cunliang_pivot = pd.concat([df_cunliang_pivot, total_df])


# In[68]:


with pd.ExcelWriter('同环比分表_%s.xlsx'%now, mode='a',engine="openpyxl") as writer:
    df_cunliang_pivot.to_excel(writer, sheet_name='省经理存量店铺环比')


# In[69]:


#省经理-存量店铺-环比
values = ['本期是否营业','环比期是否营业','本期实收金额','环比期实收金额','本期堂食实收','环比期堂食实收',
                                                '本期外卖实收','环比期外卖实收','本期自提实收','环比期自提实收']
          
df_cunliang_pivot = df_cunliang.pivot_table(index=['区域经理'],values = values,aggfunc=np.sum)


# In[70]:


df_cunliang_pivot['店铺增减'] = df_cunliang_pivot['本期是否营业']-df_cunliang_pivot['环比期是否营业']
#实收
df_cunliang_pivot['实收环比'] = (df_cunliang_pivot['本期实收金额'] - df_cunliang_pivot['环比期实收金额'])/df_cunliang_pivot['环比期实收金额']
df_cunliang_pivot.sort_values("实收环比",inplace=True,ascending=False) #排序
#堂食实收
df_cunliang_pivot['堂食实收环比'] = (df_cunliang_pivot['本期堂食实收'] - df_cunliang_pivot['环比期堂食实收'])/df_cunliang_pivot['环比期堂食实收']
#外卖实收
df_cunliang_pivot['外卖实收环比'] = (df_cunliang_pivot['本期外卖实收'] - df_cunliang_pivot['环比期外卖实收'])/df_cunliang_pivot['环比期外卖实收']
#外卖实收
df_cunliang_pivot['自提实收环比'] = (df_cunliang_pivot['本期自提实收'] - df_cunliang_pivot['环比期自提实收'])/df_cunliang_pivot['环比期自提实收']


# In[71]:


values = ['本期是否营业','环比期是否营业','店铺增减','本期实收金额','环比期实收金额','实收环比','本期堂食实收','环比期堂食实收',
          '堂食实收环比','本期外卖实收','环比期外卖实收','外卖实收环比','本期自提实收','环比期自提实收','自提实收环比']
df_cunliang_pivot = df_cunliang_pivot.reindex(columns=values)
df_cunliang_pivot.rename(columns={'本期是否营业': '本期店铺数','环比期是否营业': '环比期店铺数'}, inplace=True)


# In[72]:


total = df_cunliang_pivot[['本期店铺数', '环比期店铺数','店铺增减','本期实收金额','环比期实收金额',
            '本期堂食实收','环比期堂食实收','本期外卖实收','环比期外卖实收','本期自提实收','环比期自提实收']].sum()
total['实收环比'] = Tonghuanbi(total['本期实收金额'] , total['环比期实收金额'])

total['堂食实收环比'] = Tonghuanbi(total['本期堂食实收'] , total['环比期堂食实收'])
total['外卖实收环比'] = Tonghuanbi(total['本期外卖实收'] , total['环比期外卖实收'])
total['自提实收环比'] = Tonghuanbi(total['本期自提实收'] , total['环比期自提实收'])
total_df = pd.DataFrame(total).T.rename(index={0: '合计'})
df_cunliang_pivot = pd.concat([df_cunliang_pivot, total_df])


# In[73]:


with pd.ExcelWriter('同环比分表_%s.xlsx'%now, mode='a',engine="openpyxl") as writer:
    df_cunliang_pivot.to_excel(writer, sheet_name='区域经理存量店铺环比')


# 以下是整合表↓

# In[74]:


#全量-同比-透视
values = ['本期是否营业','同比期是否营业','本期实收金额','同比期实收金额']
df_toushi_quanliang = df_zongbiao.pivot_table(index=['大区经理','省经理','区域经理'],values = values,aggfunc=np.sum)
df_toushi_quanliang['店铺增减'] = df_toushi_quanliang['本期是否营业']-df_toushi_quanliang['同比期是否营业']
df_toushi_quanliang['实收同比'] = (df_toushi_quanliang['本期实收金额'] - df_toushi_quanliang['同比期实收金额'])/df_toushi_quanliang['同比期实收金额']
df_toushi_quanliang.rename(columns={'本期是否营业': '本期店铺数','同比期是否营业': '同比期店铺数'}, inplace=True)
columns = ['本期店铺数','同比期店铺数','店铺增减','本期实收金额','同比期实收金额','实收同比']
quanliang_tongbi = df_toushi_quanliang.reindex(columns=columns)


# In[75]:


#全量-环比-透视
values = ['本期是否营业','环比期是否营业','本期实收金额','环比期实收金额']
df_toushi_quanliang = df_zongbiao.pivot_table(index=['大区经理','省经理','区域经理'],values = values,aggfunc=np.sum)
df_toushi_quanliang['店铺增减'] = df_toushi_quanliang['本期是否营业']-df_toushi_quanliang['环比期是否营业']
df_toushi_quanliang['实收环比'] = (df_toushi_quanliang['本期实收金额'] - df_toushi_quanliang['环比期实收金额'])/df_toushi_quanliang['环比期实收金额']
df_toushi_quanliang.rename(columns={'本期是否营业': '本期店铺数','环比期是否营业': '环比期店铺数'}, inplace=True)
columns = ['本期店铺数','环比期店铺数','店铺增减','本期实收金额','环比期实收金额','实收环比']
quanliang_huanbi = df_toushi_quanliang.reindex(columns=columns)


# In[76]:


#同比期存量店铺
df_cunliang_tongbi = df_zongbiao[(df_zongbiao['本期是否营业']==1)]
df_cunliang_tongbi = df_cunliang_tongbi[(df_cunliang_tongbi['同比期是否营业']==1)]


# In[77]:


#存量-同比-透视
values = ['本期是否营业','同比期是否营业','本期实收金额','同比期实收金额']
df_toushi_quanliang = df_cunliang_tongbi.pivot_table(index=['大区经理','省经理','区域经理'],values = values,aggfunc=np.sum)
df_toushi_quanliang['店铺增减'] = df_toushi_quanliang['本期是否营业']-df_toushi_quanliang['同比期是否营业']
df_toushi_quanliang['实收同比'] = (df_toushi_quanliang['本期实收金额'] - df_toushi_quanliang['同比期实收金额'])/df_toushi_quanliang['同比期实收金额']
df_toushi_quanliang.rename(columns={'本期是否营业': '本期店铺数','同比期是否营业': '同比期店铺数'}, inplace=True)
columns = ['本期店铺数','同比期店铺数','店铺增减','本期实收金额','同比期实收金额','实收同比']
cunliang_tongbi = df_toushi_quanliang.reindex(columns=columns)


# In[78]:


#环比期存量店铺
df_cunliang_tongbi = df_zongbiao[(df_zongbiao['本期是否营业']==1)]
df_cunliang_tongbi = df_cunliang_tongbi[(df_cunliang_tongbi['环比期是否营业']==1)]


# In[79]:


#存量-环比-透视

values = ['本期是否营业','环比期是否营业','本期实收金额','环比期实收金额']
df_toushi_quanliang = df_cunliang_tongbi.pivot_table(index=['大区经理','省经理','区域经理'],values = values,aggfunc=np.sum)
df_toushi_quanliang['店铺增减'] = df_toushi_quanliang['本期是否营业']-df_toushi_quanliang['环比期是否营业']
df_toushi_quanliang['实收环比'] = (df_toushi_quanliang['本期实收金额'] - df_toushi_quanliang['环比期实收金额'])/df_toushi_quanliang['环比期实收金额']
df_toushi_quanliang.rename(columns={'本期是否营业': '本期店铺数','环比期是否营业': '环比期店铺数'}, inplace=True)
columns = ['本期店铺数','环比期店铺数','店铺增减','本期实收金额','环比期实收金额','实收环比']
cunliang_huanbi = df_toushi_quanliang.reindex(columns=columns)


# In[80]:


tongbi_weidu = pd.merge(quanliang_tongbi,cunliang_tongbi,on=['大区经理','省经理','区域经理'], how='outer')


# In[81]:


huanbi_weidu = pd.merge(quanliang_huanbi,cunliang_huanbi,on=['大区经理','省经理','区域经理'], how='outer')


# In[82]:


zong_weidu = pd.merge(tongbi_weidu,huanbi_weidu,on=['大区经理','省经理','区域经理'])


# In[83]:


zong_weidu.drop(columns=['同比期店铺数_y','店铺增减_y_x','环比期店铺数_y','店铺增减_y_y'],inplace=True) 


# In[84]:


zong_weidu.rename(columns={'本期是否营业': '本期店铺数','同比期是否营业': '同比期店铺数'}, inplace=True)


# In[85]:


with pd.ExcelWriter('同环比分表_%s.xlsx'%now, mode='a',engine="openpyxl") as writer:
    zong_weidu.to_excel(writer, sheet_name='总表',merge_cells=False)


# 以下是VBA代码：
# 1. 全表操作
# 
# ```vb
# Sub 实收金额转为万()
#     Dim i As Integer
#     Dim LastColumn As Integer
#     Dim rng As Range
#     Dim ws As Worksheet
#     
#     For Each ws In ActiveWorkbook.Worksheets
#         LastColumn = ws.Cells(1, Columns.Count).End(xlToLeft).Column
#         
#         For i = 1 To LastColumn
#             If InStr(1, ws.Cells(1, i), "期") > 0 And InStr(1, ws.Cells(1, i), "实收") > 0 Then
#                 Set rng = ws.Range(ws.Cells(2, i), ws.Cells(Rows.Count, i).End(xlUp))
#                 rng.NumberFormat = "0"".""0,""万"""
#             End If
#         Next i
#     Next ws
# End Sub
# 
# ```
# 
# 2. 透视表复制表操作（填充大区经理、省经理颜色）
# ```vb
# Sub FillColor()
#     Dim lastRow As Long
#     Dim i As Long
#     
#     '获取最后一行
#     lastRow = ActiveSheet.Cells(Rows.Count, 1).End(xlUp).Row
#     
#     '循环每一行
#     For i = 1 To lastRow
#         '如果A列单元格包含“汇总”，则填充颜色RGB(255,255,0)
#         If InStr(1, Cells(i, 1), "汇总") > 0 Then
#             Range(Cells(i, 1), Cells(i, Columns.Count)).Interior.Color = RGB(255, 255, 0)
#         End If
#         
#         '如果B列单元格包含“汇总”，则填充颜色RGB(255,192,0)
#         If InStr(1, Cells(i, 2), "汇总") > 0 Then
#             Range(Cells(i, 2), Cells(i, Columns.Count)).Interior.Color = RGB(255, 192, 0)
#         End If
#     Next i
# End Sub
# ```
# 
# 
# 

# 以下将区域经理区分出省经理

# In[86]:


filePath = '同环比分表_%s.xlsx'%now


# In[87]:


#省经理
df_sheng = pd.read_excel(filePath,sheet_name="省经理全量店铺同比")


# In[89]:


shengjingli_list = list(df_sheng['Unnamed: 0'])
#shengjingli_list = list(df_sheng.index)
shengjingli_list.remove('本期未营业')
shengjingli_list.remove('已解约')


# In[90]:


df_quyu = pd.read_excel(filePath,sheet_name="区域经理全量店铺同比", index_col=0)
df_quyu = df_quyu.iloc[:-1] #删除合计行


# In[91]:


#省经理代管
#df_quyu_sheng = df_quyu[(df_quyu['Unnamed: 0'].isin(shengjingli_list))]
df_quyu_sheng = df_quyu[df_quyu.index.isin(shengjingli_list)]
total = df_quyu_sheng[['本期店铺数', '同比期店铺数','店铺增减','本期实收金额','同比期实收金额',
            '本期堂食实收','同比期堂食实收','本期外卖实收','同比期外卖实收','本期自提实收','同比期自提实收']].sum()
total['实收同比'] = Tonghuanbi(total['本期实收金额'] , total['同比期实收金额'])
total['堂食实收同比'] = Tonghuanbi(total['本期堂食实收'] , total['同比期堂食实收'])
total['外卖实收同比'] = Tonghuanbi(total['本期外卖实收'] , total['同比期外卖实收'])
total['自提实收同比'] = Tonghuanbi(total['本期自提实收'] , total['同比期自提实收'])
total_df = pd.DataFrame(total).T.rename(index={0: '合计'})
df_quyu_sheng = pd.concat([df_quyu_sheng, total_df])
#区域经理
#df_quyu_quyu = df_quyu[~(df_quyu['Unnamed: 0'].isin(shengjingli_list))]
df_quyu_quyu = df_quyu[~df_quyu.index.isin(shengjingli_list)]
total = df_quyu_quyu[['本期店铺数', '同比期店铺数','店铺增减','本期实收金额','同比期实收金额',
            '本期堂食实收','同比期堂食实收','本期外卖实收','同比期外卖实收','本期自提实收','同比期自提实收']].sum()
total['实收同比'] = Tonghuanbi(total['本期实收金额'] , total['同比期实收金额'])
total['堂食实收同比'] = Tonghuanbi(total['本期堂食实收'] , total['同比期堂食实收'])
total['外卖实收同比'] = Tonghuanbi(total['本期外卖实收'] , total['同比期外卖实收'])
total['自提实收同比'] = Tonghuanbi(total['本期自提实收'] , total['同比期自提实收'])
total_df = pd.DataFrame(total).T.rename(index={0: '合计'})
df_quyu_quyu = pd.concat([df_quyu_quyu, total_df])


# In[92]:


with pd.ExcelWriter(filePath, mode='a',engine="openpyxl") as writer:
    df_quyu_sheng.to_excel(writer, sheet_name='区域经理全量店铺同比(省代)',index = True)
    df_quyu_quyu.to_excel(writer, sheet_name='区域经理全量店铺同比(区域)',index = True)


# In[93]:


df_quyu = pd.read_excel(filePath,sheet_name="区域经理全量店铺环比", index_col=0)
df_quyu = df_quyu.iloc[:-1] #删除合计行


# In[94]:


#省代
#df_quyu_sheng = df_quyu[(df_quyu['Unnamed: 0'].isin(shengjingli_list))]
df_quyu_sheng = df_quyu[df_quyu.index.isin(shengjingli_list)]

total = df_quyu_sheng[['本期店铺数', '环比期店铺数','店铺增减','本期实收金额','环比期实收金额',
            '本期堂食实收','环比期堂食实收','本期外卖实收','环比期外卖实收','本期自提实收','环比期自提实收']].sum()
total['实收环比'] = Tonghuanbi(total['本期实收金额'] , total['环比期实收金额'])
total['堂食实收环比'] = Tonghuanbi(total['本期堂食实收'] , total['环比期堂食实收'])
total['外卖实收环比'] = Tonghuanbi(total['本期外卖实收'] , total['环比期外卖实收'])
total['自提实收环比'] = Tonghuanbi(total['本期自提实收'] , total['环比期自提实收'])
total_df = pd.DataFrame(total).T.rename(index={0: '合计'})
df_quyu_sheng = pd.concat([df_quyu_sheng, total_df])
#区域经理（区域）
#df_quyu_quyu = df_quyu[~(df_quyu['Unnamed: 0'].isin(shengjingli_list))]
df_quyu_quyu = df_quyu[~df_quyu.index.isin(shengjingli_list)]
total = df_quyu_quyu[['本期店铺数', '环比期店铺数','店铺增减','本期实收金额','环比期实收金额',
            '本期堂食实收','环比期堂食实收','本期外卖实收','环比期外卖实收','本期自提实收','环比期自提实收']].sum()
total['实收环比'] = Tonghuanbi(total['本期实收金额'] , total['环比期实收金额'])
total['堂食实收环比'] = Tonghuanbi(total['本期堂食实收'] , total['环比期堂食实收'])
total['外卖实收环比'] = Tonghuanbi(total['本期外卖实收'] , total['环比期外卖实收'])
total['自提实收环比'] = Tonghuanbi(total['本期自提实收'] , total['环比期自提实收'])
total_df = pd.DataFrame(total).T.rename(index={0: '合计'})
df_quyu_quyu = pd.concat([df_quyu_quyu, total_df])

with pd.ExcelWriter(filePath, mode='a',engine="openpyxl") as writer:
    df_quyu_sheng.to_excel(writer, sheet_name='区域经理全量店铺环比(省代)',index = True)
    df_quyu_quyu.to_excel(writer, sheet_name='区域经理全量店铺环比(区域)',index = True)


# In[95]:


df_quyu = pd.read_excel(filePath,sheet_name="区域经理存量店铺同比", index_col=0)
df_quyu = df_quyu.iloc[:-1] #删除合计行


# In[96]:


#省经理代管
#df_quyu_sheng = df_quyu[(df_quyu['Unnamed: 0'].isin(shengjingli_list))]
df_quyu_sheng = df_quyu[df_quyu.index.isin(shengjingli_list)]
total = df_quyu_sheng[['本期店铺数', '同比期店铺数','店铺增减','本期实收金额','同比期实收金额',
            '本期堂食实收','同比期堂食实收','本期外卖实收','同比期外卖实收','本期自提实收','同比期自提实收']].sum()
total['实收同比'] = Tonghuanbi(total['本期实收金额'] , total['同比期实收金额'])
total['堂食实收同比'] = Tonghuanbi(total['本期堂食实收'] , total['同比期堂食实收'])
total['外卖实收同比'] = Tonghuanbi(total['本期外卖实收'] , total['同比期外卖实收'])
total['自提实收同比'] = Tonghuanbi(total['本期自提实收'] , total['同比期自提实收'])
total_df = pd.DataFrame(total).T.rename(index={0: '合计'})
df_quyu_sheng = pd.concat([df_quyu_sheng, total_df])
#区域经理
#df_quyu_quyu = df_quyu[~(df_quyu['Unnamed: 0'].isin(shengjingli_list))]
df_quyu_quyu = df_quyu[~df_quyu.index.isin(shengjingli_list)]
total = df_quyu_quyu[['本期店铺数', '同比期店铺数','店铺增减','本期实收金额','同比期实收金额',
            '本期堂食实收','同比期堂食实收','本期外卖实收','同比期外卖实收','本期自提实收','同比期自提实收']].sum()
total['实收同比'] = Tonghuanbi(total['本期实收金额'] , total['同比期实收金额'])
total['堂食实收同比'] = Tonghuanbi(total['本期堂食实收'] , total['同比期堂食实收'])
total['外卖实收同比'] = Tonghuanbi(total['本期外卖实收'] , total['同比期外卖实收'])
total['自提实收同比'] = Tonghuanbi(total['本期自提实收'] , total['同比期自提实收'])
total_df = pd.DataFrame(total).T.rename(index={0: '合计'})
df_quyu_quyu = pd.concat([df_quyu_quyu, total_df])

with pd.ExcelWriter(filePath, mode='a',engine="openpyxl") as writer:
    df_quyu_sheng.to_excel(writer, sheet_name='区域经理存量店铺同比(省代)',index = True)
    df_quyu_quyu.to_excel(writer, sheet_name='区域经理存量店铺同比(区域)',index = True)


# In[97]:


df_quyu = pd.read_excel(filePath,sheet_name="区域经理存量店铺环比", index_col=0)
df_quyu = df_quyu.iloc[:-1] #删除合计行


# In[98]:


#df_quyu_sheng = df_quyu[(df_quyu['Unnamed: 0'].isin(shengjingli_list))]
df_quyu_sheng = df_quyu[df_quyu.index.isin(shengjingli_list)]
total = df_quyu_sheng[['本期店铺数', '环比期店铺数','店铺增减','本期实收金额','环比期实收金额',
            '本期堂食实收','环比期堂食实收','本期外卖实收','环比期外卖实收','本期自提实收','环比期自提实收']].sum()
total['实收环比'] = Tonghuanbi(total['本期实收金额'] , total['环比期实收金额'])
total['堂食实收环比'] = Tonghuanbi(total['本期堂食实收'] , total['环比期堂食实收'])
total['外卖实收环比'] = Tonghuanbi(total['本期外卖实收'] , total['环比期外卖实收'])
total['自提实收环比'] = Tonghuanbi(total['本期自提实收'] , total['环比期自提实收'])
total_df = pd.DataFrame(total).T.rename(index={0: '合计'})
df_quyu_sheng = pd.concat([df_quyu_sheng, total_df])
#区域经理（区域）
#df_quyu_quyu = df_quyu[~(df_quyu['Unnamed: 0'].isin(shengjingli_list))]
#df_quyu_quyu = df_quyu[~(df_quyu['Unnamed: 0'].isin(shengjingli_list))]
df_quyu_quyu = df_quyu[~df_quyu.index.isin(shengjingli_list)]
total = df_quyu_quyu[['本期店铺数', '环比期店铺数','店铺增减','本期实收金额','环比期实收金额',
            '本期堂食实收','环比期堂食实收','本期外卖实收','环比期外卖实收','本期自提实收','环比期自提实收']].sum()
total['实收环比'] = Tonghuanbi(total['本期实收金额'] , total['环比期实收金额'])
total['堂食实收环比'] = Tonghuanbi(total['本期堂食实收'] , total['环比期堂食实收'])
total['外卖实收环比'] = Tonghuanbi(total['本期外卖实收'] , total['环比期外卖实收'])
total['自提实收环比'] = Tonghuanbi(total['本期自提实收'] , total['环比期自提实收'])
total_df = pd.DataFrame(total).T.rename(index={0: '合计'})
df_quyu_quyu = pd.concat([df_quyu_quyu, total_df])

with pd.ExcelWriter(filePath, mode='a',engine="openpyxl") as writer:
    df_quyu_sheng.to_excel(writer, sheet_name='区域经理存量店铺环比(省代)',index = True)
    df_quyu_quyu.to_excel(writer, sheet_name='区域经理存量店铺环比(区域)',index = True)


# In[99]:


df_quyu_quyu


# 以下合并同环比分表

# In[100]:


import openpyxl,time


# In[101]:


#同环比分表
#file = r'C:\Users\admin\OneDrive\甜啦啦\同环比分表.xlsx'
file = '同环比分表_%s.xlsx'%now
file2 = r'C:\Users\admin\OneDrive\甜啦啦\甜啦啦代码\代码底表\大区省区域经理分表模板.xlsx'
#file2 = r'C:\Users\admin\OneDrive\甜啦啦\甜啦啦代码\代码底表\大区省区域经理分表模板_bak.xlsx'


# In[102]:


source_file = openpyxl.load_workbook(file)
target_file = openpyxl.load_workbook(file2)


# In[103]:


#大区经理
target_sheet = target_file["大区经理排名"]

source_sheet = source_file["大区经理全量店铺同比"]

for r in range(1, source_sheet.max_row + 1):
    for c in range(1, source_sheet.max_column + 1):
        target_sheet.cell(row=r+2, column=c+2).value = source_sheet.cell(row=r, column=c).value

source_sheet = source_file["大区经理全量店铺环比"]

for r in range(1, source_sheet.max_row + 1):
    for c in range(1, source_sheet.max_column + 1):
        target_sheet.cell(row=r+2, column=c+21).value = source_sheet.cell(row=r, column=c).value

source_sheet = source_file["大区经理存量店铺同比"]

for r in range(1, source_sheet.max_row + 1):
    for c in range(1, source_sheet.max_column + 1):
        target_sheet.cell(row=r+34, column=c+2).value = source_sheet.cell(row=r, column=c).value

source_sheet = source_file["大区经理存量店铺环比"]

for r in range(1, source_sheet.max_row + 1):
    for c in range(1, source_sheet.max_column + 1):
        target_sheet.cell(row=r+34, column=c+21).value = source_sheet.cell(row=r, column=c).value

target_file.save("大区省区域经理分表输出.xlsx")


# In[104]:


#省经理
target_sheet = target_file["省经理排名"]

source_sheet = source_file["省经理全量店铺同比"]

for r in range(1, source_sheet.max_row + 1):
    for c in range(1, source_sheet.max_column + 1):
        target_sheet.cell(row=r+2, column=c+2).value = source_sheet.cell(row=r, column=c).value

source_sheet = source_file["省经理全量店铺环比"]

for r in range(1, source_sheet.max_row + 1):
    for c in range(1, source_sheet.max_column + 1):
        target_sheet.cell(row=r+2, column=c+21).value = source_sheet.cell(row=r, column=c).value

source_sheet = source_file["省经理存量店铺同比"]

for r in range(1, source_sheet.max_row + 1):
    for c in range(1, source_sheet.max_column + 1):
        target_sheet.cell(row=r+58, column=c+2).value = source_sheet.cell(row=r, column=c).value

source_sheet = source_file["省经理存量店铺环比"]

for r in range(1, source_sheet.max_row + 1):
    for c in range(1, source_sheet.max_column + 1):
        target_sheet.cell(row=r+58, column=c+21).value = source_sheet.cell(row=r, column=c).value

#target_file.save("大区省区域经理分表输出.xlsx")


# In[105]:


#区域经理(区域)
target_sheet = target_file["区域经理排名"]

source_sheet = source_file["区域经理全量店铺同比(区域)"]

for r in range(1, source_sheet.max_row + 1):
    for c in range(1, source_sheet.max_column + 1):
        target_sheet.cell(row=r+2, column=c+2).value = source_sheet.cell(row=r, column=c).value

source_sheet = source_file["区域经理全量店铺环比(区域)"]

for r in range(1, source_sheet.max_row + 1):
    for c in range(1, source_sheet.max_column + 1):
        target_sheet.cell(row=r+2, column=c+21).value = source_sheet.cell(row=r, column=c).value

source_sheet = source_file["区域经理存量店铺同比(区域)"]

for r in range(1, source_sheet.max_row + 1):
    for c in range(1, source_sheet.max_column + 1):
        target_sheet.cell(row=r+143, column=c+2).value = source_sheet.cell(row=r, column=c).value

source_sheet = source_file["区域经理存量店铺环比(区域)"]

for r in range(1, source_sheet.max_row + 1):
    for c in range(1, source_sheet.max_column + 1):
        target_sheet.cell(row=r+143, column=c+21).value = source_sheet.cell(row=r, column=c).value

#区域经理(省代)        
source_sheet = source_file["区域经理全量店铺同比(省代)"]

for r in range(1, source_sheet.max_row + 1):
    for c in range(1, source_sheet.max_column + 1):
        target_sheet.cell(row=r+283, column=c+2).value = source_sheet.cell(row=r, column=c).value

source_sheet = source_file["区域经理全量店铺环比(省代)"]

for r in range(1, source_sheet.max_row + 1):
    for c in range(1, source_sheet.max_column + 1):
        target_sheet.cell(row=r+283, column=c+21).value = source_sheet.cell(row=r, column=c).value

source_sheet = source_file["区域经理存量店铺同比(省代)"]

for r in range(1, source_sheet.max_row + 1):
    for c in range(1, source_sheet.max_column + 1):
        target_sheet.cell(row=r+334, column=c+2).value = source_sheet.cell(row=r, column=c).value

source_sheet = source_file["区域经理存量店铺环比(省代)"]

for r in range(1, source_sheet.max_row + 1):
    for c in range(1, source_sheet.max_column + 21):
        target_sheet.cell(row=r+334, column=c+21).value = source_sheet.cell(row=r, column=c).value        
        
        
        
target_file.save("大区省区域经理分表输出_%s.xlsx"%now)


# In[106]:


"大区省区域经理分表输出_%s.xlsx"%now


# In[107]:


file


# Sub ConvertTextToNumber()
#     Dim rng As Range
#     Dim col As Variant
#     '定义一个数组，存储要转换的列号
#     Dim colsArray(1 To 3) As Variant
#     colsArray(1) = "I"
#     colsArray(2) = "L"
#     colsArray(3) = "O"
# 	colsArray(4) = "R"
# 	colsArray(5) = "AB"
# 	colsArray(6) = "AE"
# 	colsArray(7) = "AH"
# 	colsArray(8) = "AK"
#     '遍历数组中的每个列号
#     For Each col In colsArray
#         '获取该列的范围
#         Set rng = Range(col & "1:" & col & Rows.Count)
#         '将单元格格式改为常规
#         rng.NumberFormat = "General"
#         '使用Value属性将文本数字转换为常规数字
#         rng.Value = rng.Value
#     Next col
# End Sub
# 
# 

# 
# Sub SetConditionalFormat()
#     Dim rng As Range
#     Dim col As Variant '修改这里
#     Dim lastVal As Variant
#     '获取已有选区范围
#     Set rng = Selection
#     '定义一个数组，存储要作用的列号
#     Dim colsArray(1 To 4) As Variant '修改这里
#     colsArray(1) = 9
#     colsArray(2) = 12
#     colsArray(3) = 15
#     colsArray(4) = 18
#     '遍历数组中的每个列号
#     For Each col In colsArray
#         '获取最下方的值
#         lastVal = rng.Cells(rng.Rows.Count, col).Value
#         '如果是数字，则设置条件格式
#         If IsNumeric(lastVal) Then
#             '添加条件格式规则，如果小于最下方的值，则字体为红色
#             rng.Columns(col).FormatConditions.Add Type:=xlCellValue, Operator:=xlLess, Formula1:=lastVal
#             rng.Columns(col).FormatConditions(1).Font.Color = vbRed
#         End If
#     Next col
# End Sub
# 
# 

# In[ ]:




