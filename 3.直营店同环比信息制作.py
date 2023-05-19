#!/usr/bin/env python
# coding: utf-8

# In[1]:


from openpyxl.utils import range_boundaries
from openpyxl import load_workbook
import time
import configparser
import glob
import os
import openpyxl
import pandas as pd

conf = configparser.ConfigParser()
conf.read('config.ini', encoding="utf-8-sig")
workfloder = conf.get('Path', 'workfloder')
zhiying = conf.get('Path', 'teshushop')  # 特使店铺汇总
zhiyingdibiao = conf.get('Path', 'zhiyingdibiao')  # 特使店铺汇总
fileList = glob.glob(os.path.join(workfloder, '*.xlsx'))
tongbiyuanbiao = [file for file in fileList if os.path.basename(file).startswith('516营业同比表')][-1]
huanbiyuanbiao = [file for file in fileList if os.path.basename(file).startswith('516营业环比表')][-1]
benqishouying = [file for file in fileList if os.path.basename(file).startswith('112收银汇总表本期')][-1]
shangqishouying = [file for file in fileList if os.path.basename(file).startswith('112收银汇总表环比期')][-1]
shangqishouying = [file for file in fileList if os.path.basename(file).startswith('112收银汇总表环比期')][-1]
qunianshouying = [file for file in fileList if os.path.basename(file).startswith('112收银汇总表同比期')][-1]
print(f'检测内容如下：\n工作目录：{workfloder}\n直营店：{zhiying}\n同比表：{tongbiyuanbiao} \n环比表：{huanbiyuanbiao}\n本期收银表：{benqishouying}\n上期收银表：{shangqishouying}\n去年同期收银表：{qunianshouying}')
now = time.strftime('%Y-%m-%d')
output_file_name = workfloder + f'\\大区省区域经理分表输出_{now}.xlsx'


# In[4]:


#取消合并单元格、填充各期
def tableFormatting(wb):
    sheet = wb['Report']
    merge_list = []
    for merge in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(merge.coord)
        if min_row == 3 and max_row == 4: # 只选取第三行和第四行的合并单元格
            merge_list.append(merge)

    for group in merge_list:
        min_col, min_row, max_col, max_row = range_boundaries(group.coord)
        top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
        sheet.unmerge_cells(str(group))
        for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
            for cell in row:
                cell.value = top_left_cell_value

    for cell in sheet[4]: # 遍历第四行的单元格
        for merge in sheet.merged_cells.ranges: # 遍历所有的合并单元格
            if cell.row - 1 == merge.max_row and cell.column >= merge.min_col and cell.column <= merge.max_col and merge.min_row == merge.max_row and merge.min_row == 3: # 如果这个单元格对应的第三行单元格是左右合并的单元格
                merge_value = sheet.cell(merge.min_row,merge.min_col).value # 获取合并单元格的值
                cell.value = merge_value + '_' + cell.value # 修改这个单元格的值为合并单元格的值+"_"+它本身的值
                break # 跳出循环
    return wb


# In[5]:


#同比表格式化
wb = load_workbook(tongbiyuanbiao)
wb = tableFormatting(wb)
sheet = wb['Report']
for cell in sheet[4]: # 遍历第四行的单元格
    if '对比期' in cell.value: # 如果单元格的值包含'对比期'
        cell.value = cell.value.replace('对比期', '同比期') # 替换单元格的值中的'对比期'为'同比期'
    if '_' in cell.value: # 如果单元格的值包含'_'
        cell.value = cell.value.replace('_', '\n') # 替换单元格的值中的'_'为换行符
        
wb.save('同比表格式化.xlsx')


# In[6]:


#同比表格式化
wb = load_workbook(huanbiyuanbiao)
wb = tableFormatting(wb)
sheet = wb['Report']
for cell in sheet[4]: # 遍历第四行的单元格
    if '对比期' in cell.value: # 如果单元格的值包含'对比期'
        cell.value = cell.value.replace('对比期', '环比期') # 替换单元格的值中的'对比期'为'同比期'
    if '_' in cell.value: # 如果单元格的值包含'_'
        cell.value = cell.value.replace('_', '\n') # 替换单元格的值中的'_'为换行符
        
wb.save('环比表格式化.xlsx')


# #### 收银

# In[7]:



def shouyinFormatting(wb):
    sheet = wb['Report']
    merge_list = []
    for merge in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(merge.coord)
        if min_row == 3 and max_row == 4: # 只选取第三行和第四行的合并单元格
            merge_list.append(merge)
    for group in merge_list:
        min_col, min_row, max_col, max_row = range_boundaries(group.coord)
        top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
        sheet.unmerge_cells(str(group))
        for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
            for cell in row:
                cell.value = top_left_cell_value
    return wb


# In[8]:


#本期收银格式化
wb = load_workbook(benqishouying)
wb = shouyinFormatting(wb)
sheet = wb['Report']
for cell in sheet[4]: # 第四行的所有单元格
    if cell.value: # 如果单元格不为空
        cell.value = cell.value+"\n本期"  # 在原本值前加上“本期”和换行符
    
wb.save('本期收银表格式化.xlsx')


# In[9]:


#环比期收银格式化
wb = load_workbook(shangqishouying)
wb = shouyinFormatting(wb)
sheet = wb['Report']
for cell in sheet[4]: # 第四行的所有单元格
    if cell.value: # 如果单元格不为空
        cell.value = cell.value+"\n环比期"  # 在原本值前加上“本期”和换行符
    
wb.save('环比期收银表格式化.xlsx')


# In[10]:


#同比期收银格式化
wb = load_workbook(qunianshouying)
wb = shouyinFormatting(wb)
sheet = wb['Report']
for cell in sheet[4]: # 第四行的所有单元格
    if cell.value: # 如果单元格不为空
        cell.value = cell.value+"\n同比期"  # 在原本值前加上“本期”和换行符
    
wb.save('同比期收银表格式化.xlsx')


# In[11]:


import pandas as pd
import numpy as np


# In[12]:


huanbigeshi_df = pd.read_excel('环比表格式化.xlsx',sheet_name=0,header = 3)
tongbigeshi_df = pd.read_excel('同比表格式化.xlsx',sheet_name=0,header = 3)
benqishouying_df = pd.read_excel('本期收银表格式化.xlsx',sheet_name=0,header = 3)
huanbishouying_df = pd.read_excel('环比期收银表格式化.xlsx',sheet_name=0,header = 3)
tongbishouying_df = pd.read_excel('同比期收银表格式化.xlsx',sheet_name=0,header = 3)


# In[13]:


huanbigeshi_df.shape 


# In[14]:


tongbigeshi_df.shape 


# In[15]:


benqishouying_df.shape


# In[16]:


huanbishouying_df.shape


# In[17]:


tongbishouying_df.shape


# 环比表

# In[18]:


huanbi_df = pd.merge(huanbigeshi_df,benqishouying_df,left_on ='店铺名称',right_on ='店铺名称\n本期', how='outer')
huanbi_df = pd.merge(huanbi_df,huanbishouying_df,left_on ='店铺名称',right_on ='店铺名称\n环比期', how='outer')


# In[19]:


huanbi_df.shape 


# In[20]:


huanbi_df.rename(columns={
    '营业天数\n本期_x': '营业天数\n本期','营业天数\n环比期_x': '营业天数\n环比期',
    '流水金额\n本期_x': '流水金额\n本期','流水金额\n环比期_x': '流水金额\n环比期',
    '实收金额\n本期_x': '实收金额\n本期','实收金额\n环比期_x': '实收金额\n环比期',
    '账单数\n本期_x': '账单数\n本期','账单数\n环比期_x': '账单数\n环比期',
    '单均消费\n本期_x': '单均消费\n本期','单均消费\n环比期_x': '单均消费\n环比期',
    '流水金额\n增长%': '流水金额\n环比',
    '实收金额\n增长%': '实收金额\n环比',
    '自提流水\n增长%': '自提流水\n环比',
    '自提实收\n本期_x': '自提实收\n本期','自提实收\n环比期_x': '自提实收\n环比期',
}, inplace=True)


# In[21]:


huanbi_df['营业天数\n差额'] = huanbi_df['营业天数\n本期'] -  huanbi_df['营业天数\n环比期']
huanbi_df['实收率\n本期'] =huanbi_df['实收金额\n本期']/ huanbi_df['流水金额\n本期']
huanbi_df['实收率\n环比期'] =  huanbi_df['实收金额\n环比期']/huanbi_df['流水金额\n环比期']
huanbi_df['实收率\n差额'] = huanbi_df['实收率\n本期'] -  huanbi_df['实收率\n环比期']
huanbi_df['账单数\n差额'] = huanbi_df['账单数\n本期'] -  huanbi_df['账单数\n环比期']
huanbi_df['单均消费\n差额'] = huanbi_df['单均消费\n本期'] -  huanbi_df['单均消费\n环比期']
huanbi_df['堂食流水\n环比'] =( huanbi_df['堂食流水\n本期'] -  huanbi_df['堂食流水\n环比期'])/huanbi_df['堂食流水\n环比期']
huanbi_df['堂食实收\n环比'] =( huanbi_df['堂食实收\n本期'] -  huanbi_df['堂食实收\n环比期'])/huanbi_df['堂食实收\n环比期']
huanbi_df['堂食实收率\n本期'] =huanbi_df['堂食实收\n本期']/ huanbi_df['堂食流水\n本期']
huanbi_df['堂食实收率\n环比期'] =  huanbi_df['堂食实收\n环比期']/huanbi_df['堂食流水\n环比期']
huanbi_df['堂食实收率\n差额'] = huanbi_df['堂食实收率\n本期'] -  huanbi_df['堂食实收率\n环比期']
huanbi_df['堂食单数\n环比'] =( huanbi_df['堂食单数\n本期'] -  huanbi_df['堂食单数\n环比期'])/huanbi_df['堂食单数\n环比期']
huanbi_df['外卖流水\n环比'] =( huanbi_df['外卖流水\n本期'] -  huanbi_df['外卖流水\n环比期'])/huanbi_df['外卖流水\n环比期']
huanbi_df['外卖实收\n环比'] =( huanbi_df['外卖实收\n本期'] -  huanbi_df['外卖实收\n环比期'])/huanbi_df['外卖实收\n环比期']
huanbi_df['外卖实收率\n本期'] =huanbi_df['外卖实收\n本期']/ huanbi_df['外卖流水\n本期']
huanbi_df['外卖实收率\n环比期'] =  huanbi_df['外卖实收\n环比期']/huanbi_df['外卖流水\n环比期']
huanbi_df['外卖实收率\n差额'] = huanbi_df['外卖实收率\n本期'] -  huanbi_df['外卖实收率\n环比期']
huanbi_df['自提流水\n环比'] =( huanbi_df['自提流水\n本期'] -  huanbi_df['自提流水\n环比期'])/huanbi_df['自提流水\n环比期']
huanbi_df['自提实收\n环比'] =( huanbi_df['自提实收\n本期'] -  huanbi_df['自提实收\n环比期'])/huanbi_df['自提实收\n环比期']
huanbi_df['自提实收率\n本期'] =huanbi_df['自提实收\n本期']/ huanbi_df['自提流水\n本期']
huanbi_df['自提实收率\n环比期'] =  huanbi_df['自提实收\n环比期']/huanbi_df['自提流水\n环比期']
huanbi_df['自提实收率\n差额'] = huanbi_df['自提实收率\n本期'] -  huanbi_df['自提实收率\n环比期']


# In[22]:


col = ['店铺名称','营业天数\n本期','营业天数\n环比期','营业天数\n差额',
       '流水金额\n本期','流水金额\n环比期','流水金额\n环比',
       '实收金额\n本期','实收金额\n环比期','实收金额\n环比','实收率\n本期','实收率\n环比期','实收率\n差额',
       '账单数\n本期','账单数\n环比期','账单数\n差额',
       '单均消费\n本期','单均消费\n环比期','单均消费\n差额',
       '堂食流水\n本期','堂食流水\n环比期','堂食流水\n环比',
       '堂食实收\n本期','堂食实收\n环比期','堂食实收\n环比',
       '堂食实收率\n本期','堂食实收率\n环比期','堂食实收率\n差额','堂食单数\n本期','堂食单数\n环比期','堂食单数\n环比',
       '外卖流水\n本期','外卖流水\n环比期','外卖流水\n环比','外卖实收\n本期','外卖实收\n环比期','外卖实收\n环比',
       '外卖实收率\n本期','外卖实收率\n环比期','外卖实收率\n差额','自提流水\n本期','自提流水\n环比期','自提流水\n环比',
       '自提实收\n本期','自提实收\n环比期','自提实收\n环比','自提实收率\n本期','自提实收率\n环比期','自提实收率\n差额',
       
      ]


# In[23]:


#设置value的显示长度为200，默认为50
#pd.set_option('max_colwidth',200)
#显示所有列，把行显示设置成最大
pd.set_option('display.max_columns', None)
#显示所有行，把列显示设置成最大
#pd.set_option('display.max_rows', None)


# In[24]:


result_huanbi_df = huanbi_df.reindex(columns=col)


# In[25]:


result_huanbi_df.head(1)


# In[26]:


result_huanbi_df.to_excel('同环比表总表_%s.xlsx'%now,merge_cells=False,sheet_name='环比表',index=False)


# In[27]:


zhiying_df = pd.read_excel(zhiying,sheet_name='直营店',header = 0)
dianpu_name = zhiying_df['哗啦啦店铺名称'].tolist()
result_zhiying_huanbi_df = result_huanbi_df[(result_huanbi_df['店铺名称'].isin(dianpu_name))]
#result_df = yuanshi_groupby.loc[zhangdanNo]


# In[28]:


result_zhiying_huanbi_df.shape 


# In[29]:


result_zhiying_huanbi_df.to_excel('直营店同环比表_%s.xlsx'%now,merge_cells=False,sheet_name='环比表',index=False)


# In[30]:


tongbigeshi_df.shape 


# 同比表

# In[31]:


tongbi_df = pd.merge(tongbigeshi_df,benqishouying_df,left_on ='店铺名称',right_on ='店铺名称\n本期', how='outer')
tongbi_df.shape


# In[32]:


tongbi_df = pd.merge(tongbi_df,tongbishouying_df,left_on ='店铺名称',right_on ='店铺名称\n同比期', how='outer')
tongbi_df.shape


# In[33]:


tongbi_df.rename(columns={
    '营业天数\n本期_x': '营业天数\n本期','营业天数\n同比期_x': '营业天数\n同比期',
    '流水金额\n本期_x': '流水金额\n本期','流水金额\n同比期_x': '流水金额\n同比期',
    '实收金额\n本期_x': '实收金额\n本期','实收金额\n同比期_x': '实收金额\n同比期',
    '账单数\n本期_x': '账单数\n本期','账单数\n同比期_x': '账单数\n同比期',
    '单均消费\n本期_x': '单均消费\n本期','单均消费\n同比期_x': '单均消费\n同比期',
    '流水金额\n增长%': '流水金额\n同比',
    '实收金额\n增长%': '实收金额\n同比',
    '自提流水\n增长%': '自提流水\n同比',
    '自提实收\n本期_x': '自提实收\n本期','自提实收\n同比期_x': '自提实收\n同比期',
}, inplace=True)


# In[34]:


tongbi_df['营业天数\n差额'] = tongbi_df['营业天数\n本期'] -  tongbi_df['营业天数\n同比期']
tongbi_df['实收率\n本期'] =tongbi_df['实收金额\n本期']/ tongbi_df['流水金额\n本期']
tongbi_df['实收率\n同比期'] =  tongbi_df['实收金额\n同比期']/tongbi_df['流水金额\n同比期']
tongbi_df['实收率\n差额'] = tongbi_df['实收率\n本期'] -  tongbi_df['实收率\n同比期']
tongbi_df['账单数\n差额'] = tongbi_df['账单数\n本期'] -  tongbi_df['账单数\n同比期']
tongbi_df['单均消费\n差额'] = tongbi_df['单均消费\n本期'] -  tongbi_df['单均消费\n同比期']
tongbi_df['堂食流水\n同比'] =( tongbi_df['堂食流水\n本期'] -  tongbi_df['堂食流水\n同比期'])/tongbi_df['堂食流水\n同比期']
tongbi_df['堂食实收\n同比'] =( tongbi_df['堂食实收\n本期'] -  tongbi_df['堂食实收\n同比期'])/tongbi_df['堂食实收\n同比期']
tongbi_df['堂食实收率\n本期'] =tongbi_df['堂食实收\n本期']/ tongbi_df['堂食流水\n本期']
tongbi_df['堂食实收率\n同比期'] =  tongbi_df['堂食实收\n同比期']/tongbi_df['堂食流水\n同比期']
tongbi_df['堂食实收率\n差额'] = tongbi_df['堂食实收率\n本期'] -  tongbi_df['堂食实收率\n同比期']
tongbi_df['堂食单数\n同比'] =( tongbi_df['堂食单数\n本期'] -  tongbi_df['堂食单数\n同比期'])/tongbi_df['堂食单数\n同比期']
tongbi_df['外卖流水\n同比'] =( tongbi_df['外卖流水\n本期'] -  tongbi_df['外卖流水\n同比期'])/tongbi_df['外卖流水\n同比期']
tongbi_df['外卖实收\n同比'] =( tongbi_df['外卖实收\n本期'] -  tongbi_df['外卖实收\n同比期'])/tongbi_df['外卖实收\n同比期']
tongbi_df['外卖实收率\n本期'] =tongbi_df['外卖实收\n本期']/ tongbi_df['外卖流水\n本期']
tongbi_df['外卖实收率\n同比期'] =  tongbi_df['外卖实收\n同比期']/tongbi_df['外卖流水\n同比期']
tongbi_df['外卖实收率\n差额'] = tongbi_df['外卖实收率\n本期'] -  tongbi_df['外卖实收率\n同比期']
tongbi_df['自提流水\n同比'] =( tongbi_df['自提流水\n本期'] -  tongbi_df['自提流水\n同比期'])/tongbi_df['自提流水\n同比期']
tongbi_df['自提实收\n同比'] =( tongbi_df['自提实收\n本期'] -  tongbi_df['自提实收\n同比期'])/tongbi_df['自提实收\n同比期']
tongbi_df['自提实收率\n本期'] =tongbi_df['自提实收\n本期']/ tongbi_df['自提流水\n本期']
tongbi_df['自提实收率\n同比期'] =  tongbi_df['自提实收\n同比期']/tongbi_df['自提流水\n同比期']
tongbi_df['自提实收率\n差额'] = tongbi_df['自提实收率\n本期'] -  tongbi_df['自提实收率\n同比期']


# In[35]:


col = ['店铺名称','营业天数\n本期','营业天数\n同比期','营业天数\n差额',
       '流水金额\n本期','流水金额\n同比期','流水金额\n同比',
       '实收金额\n本期','实收金额\n同比期','实收金额\n同比','实收率\n本期','实收率\n同比期','实收率\n差额',
       '账单数\n本期','账单数\n同比期','账单数\n差额',
       '单均消费\n本期','单均消费\n同比期','单均消费\n差额',
       '堂食流水\n本期','堂食流水\n同比期','堂食流水\n同比',
       '堂食实收\n本期','堂食实收\n同比期','堂食实收\n同比',
       '堂食实收率\n本期','堂食实收率\n同比期','堂食实收率\n差额','堂食单数\n本期','堂食单数\n同比期','堂食单数\n同比',
       '外卖流水\n本期','外卖流水\n同比期','外卖流水\n同比','外卖实收\n本期','外卖实收\n同比期','外卖实收\n同比',
       '外卖实收率\n本期','外卖实收率\n同比期','外卖实收率\n差额','自提流水\n本期','自提流水\n同比期','自提流水\n同比',
       '自提实收\n本期','自提实收\n同比期','自提实收\n同比','自提实收率\n本期','自提实收率\n同比期','自提实收率\n差额',
       
      ]


# In[36]:


tongbi_df.shape


# In[37]:


#设置value的显示长度为200，默认为50
#pd.set_option('max_colwidth',200)
#显示所有列，把行显示设置成最大
pd.set_option('display.max_columns', None)
#显示所有行，把列显示设置成最大
#pd.set_option('display.max_rows', None)


# In[38]:


result_tongbi_df = tongbi_df.reindex(columns=col)


# In[39]:


result_tongbi_df.shape


# In[40]:


result_tongbi_df.head(1)


# In[41]:


with pd.ExcelWriter('同环比表总表_%s.xlsx'%now, mode='a',engine="openpyxl") as writer:
    result_tongbi_df.to_excel(writer, sheet_name='同比表',index = False)


# In[42]:


result_tongbi_df.shape


# In[43]:


#zhiying_df = pd.read_excel(zhiying,sheet_name=0,header = 0)
dianpu_name = zhiying_df['哗啦啦店铺名称'].tolist()
result_zhiying_tongbi_df = result_tongbi_df[(result_tongbi_df['店铺名称'].isin(dianpu_name))]


# In[44]:


result_zhiying_tongbi_df.shape 


# In[45]:


with pd.ExcelWriter('直营店同环比表_%s.xlsx'%now, mode='a',engine="openpyxl") as writer:
    result_zhiying_tongbi_df.to_excel(writer, sheet_name='同比表',index = False)


# In[46]:


shanchu_df =  pd.read_excel('环比表格式化.xlsx',sheet_name=0,header = 0)


# In[47]:


#直营环比表
df =  pd.read_excel(huanbiyuanbiao,sheet_name=0,header = 2)
zhiying_huanbi_df = df[(df['店铺名称'].isin(dianpu_name))]
#直营同比表
df =  pd.read_excel(tongbiyuanbiao,sheet_name=0,header = 2)
zhiying_tongbi_df = df[(df['店铺名称'].isin(dianpu_name))]
#直营本期收银
df =  pd.read_excel(benqishouying,sheet_name=0,header = 2)
zhiying_benqishouyin_df = df[(df['店铺名称'].isin(dianpu_name))]
#直营环比期收银
df =  pd.read_excel(shangqishouying,sheet_name=0,header = 2)
zhiying_huanbishouyin_df = df[(df['店铺名称'].isin(dianpu_name))]
#直营同比期收银
df =  pd.read_excel(qunianshouying,sheet_name=0,header = 2)
zhiying_tongbishouyin_df = df[(df['店铺名称'].isin(dianpu_name))]


# In[48]:


with pd.ExcelWriter('直营店同环比表_%s.xlsx'%now, mode='a',engine="openpyxl") as writer:
    zhiying_huanbi_df.to_excel(writer, sheet_name='直营环比表',index = False)
    zhiying_tongbi_df.to_excel(writer, sheet_name='直营同比表',index = False)
    zhiying_benqishouyin_df.to_excel(writer, sheet_name='直营本期收银',index = False)
    zhiying_huanbishouyin_df.to_excel(writer, sheet_name='直营环比期收银',index = False)
    zhiying_tongbishouyin_df.to_excel(writer, sheet_name='直营同比期收银',index = False)


# In[49]:


'直营店同环比表_%s.xlsx'%now


# In[50]:


import openpyxl


# In[51]:


file = r'直营店同环比表_%s.xlsx'%now
file2 = zhiyingdibiao
#file1 = r'C:\Users\admin\OneDrive\甜啦啦\2023年04月表\直营店同环比表_2023-04.xlsx'


# In[52]:


source_file = openpyxl.load_workbook(file) #目标
target_file = openpyxl.load_workbook(file2) #模板


# In[53]:


# 选择源文件和目标文件的工作表
source_sheet = source_file['环比表']
target_sheet = target_file.active


# In[54]:


# 获取源文件的 A 列数据
source_col_A = [cell.value for cell in source_sheet['A'][1:]]

# 将 A 列数据复制到目标文件的指定区域
for i, value in enumerate(source_col_A):
    target_sheet.cell(row=4+i, column=1, value=value)


# In[55]:


#本期数据
# 获取源文件的 E 列、H 列、K 列、AF 列、AI 列和 AL 列数据
source_col_E = [cell.value for cell in source_sheet['E'][1:]]
source_col_H = [cell.value for cell in source_sheet['H'][1:]]
source_col_K = [cell.value for cell in source_sheet['K'][1:]]
source_col_AF = [cell.value for cell in source_sheet['AF'][1:]]
source_col_AI = [cell.value for cell in source_sheet['AI'][1:]]
source_col_AL = [cell.value for cell in source_sheet['AL'][1:]]

# 将 E 列数据复制到目标文件的 B 列指定区域
for i, value in enumerate(source_col_E):
    target_sheet.cell(row=4+i, column=4, value=value)

# 将 H 列数据复制到目标文件的 C 列指定区域
for i, value in enumerate(source_col_H):
    target_sheet.cell(row=4+i, column=5, value=value)

# 将 K 列数据复制到目标文件的 D 列指定区域
for i, value in enumerate(source_col_K):
    target_sheet.cell(row=4+i, column=6, value=value)

# 将 AF 列数据复制到目标文件的 E 列指定区域
for i, value in enumerate(source_col_AF):
    target_sheet.cell(row=4+i, column=7, value=value)

# 将 AI 列数据复制到目标文件的 F 列指定区域
for i, value in enumerate(source_col_AI):
    target_sheet.cell(row=4+i, column=8, value=value)

# 将 AL 列数据复制到目标文件的 G 列指定区域
for i, value in enumerate(source_col_AL):
    target_sheet.cell(row=4+i, column=9, value=value)


# In[56]:


#环比期数据
# 获取源文件的 E 列、H 列、K 列、AF 列、AI 列和 AL 列数据
source_col_E = [cell.value for cell in source_sheet['F'][1:]]
source_col_H = [cell.value for cell in source_sheet['I'][1:]]
source_col_K = [cell.value for cell in source_sheet['L'][1:]]
source_col_AF = [cell.value for cell in source_sheet['AG'][1:]]
source_col_AI = [cell.value for cell in source_sheet['AJ'][1:]]
source_col_AL = [cell.value for cell in source_sheet['AM'][1:]]

# 将 E 列数据复制到目标文件的 B 列指定区域
for i, value in enumerate(source_col_E):
    target_sheet.cell(row=4+i, column=11, value=value)

# 将 H 列数据复制到目标文件的 C 列指定区域
for i, value in enumerate(source_col_H):
    target_sheet.cell(row=4+i, column=12, value=value)

# 将 K 列数据复制到目标文件的 D 列指定区域
for i, value in enumerate(source_col_K):
    target_sheet.cell(row=4+i, column=13, value=value)

# 将 AF 列数据复制到目标文件的 E 列指定区域
for i, value in enumerate(source_col_AF):
    target_sheet.cell(row=4+i, column=14, value=value)

# 将 AI 列数据复制到目标文件的 F 列指定区域
for i, value in enumerate(source_col_AI):
    target_sheet.cell(row=4+i, column=15, value=value)

# 将 AL 列数据复制到目标文件的 G 列指定区域
for i, value in enumerate(source_col_AL):
    target_sheet.cell(row=4+i, column=16, value=value)


# In[57]:


source_sheet = source_file['同比表']
#同比期数据
# 获取源文件的 E 列、H 列、K 列、AF 列、AI 列和 AL 列数据
source_col_E = [cell.value for cell in source_sheet['F'][1:]]
source_col_H = [cell.value for cell in source_sheet['I'][1:]]
source_col_K = [cell.value for cell in source_sheet['L'][1:]]
source_col_AF = [cell.value for cell in source_sheet['AG'][1:]]
source_col_AI = [cell.value for cell in source_sheet['AJ'][1:]]
source_col_AL = [cell.value for cell in source_sheet['AM'][1:]]

# 将 E 列数据复制到目标文件的 B 列指定区域
for i, value in enumerate(source_col_E):
    target_sheet.cell(row=4+i, column=18, value=value)

# 将 H 列数据复制到目标文件的 C 列指定区域
for i, value in enumerate(source_col_H):
    target_sheet.cell(row=4+i, column=19, value=value)

# 将 K 列数据复制到目标文件的 D 列指定区域
for i, value in enumerate(source_col_K):
    target_sheet.cell(row=4+i, column=20, value=value)

# 将 AF 列数据复制到目标文件的 E 列指定区域
for i, value in enumerate(source_col_AF):
    target_sheet.cell(row=4+i, column=21, value=value)

# 将 AI 列数据复制到目标文件的 F 列指定区域
for i, value in enumerate(source_col_AI):
    target_sheet.cell(row=4+i, column=22, value=value)

# 将 AL 列数据复制到目标文件的 G 列指定区域
for i, value in enumerate(source_col_AL):
    target_sheet.cell(row=4+i, column=23, value=value)


# In[58]:


# 保存目标文件
target_file.save('直营店流水进度与同环比信息%s.xlsx'%now)


# In[59]:


'直营店流水进度与同环比信息%s.xlsx'%now


# In[ ]:




